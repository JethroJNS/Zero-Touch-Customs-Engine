from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import List, Any, Optional, Dict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import EXCEL_SHEET_ORDER, EXCEL_COLUMNS
from ml.src.extraction.merger import ShipmentEntities
from ml.src.extraction.items import ItemEntity
from ml.src.extraction import lookups as lk
from ml.src.extraction import rules as br

logger = logging.getLogger(__name__)

# NDPBM DEFAULT
DEFAULT_NDPBM: float = 2640.37

# IMPORTER DEFAULT
DEFAULT_IMPORTER = {
    "name": "PT SINAR SURYA UTAMA",
    "address": "JL. RAYA CIKARANG KAV.E-12, KAWASAN INDUSTRI MM2100, CIKARANG, BEKASI, JAWA BARAT",
    "npwp": "1000000006636570000000",
    "nib": "0311250108356000000",
    "api_type": "02",
    "kode_kantor": "040300",
    "kode_cara_bayar": "1",
}


class ExcelExporter:
    def __init__(self, template_path: Optional[str | Path] = None):
        self.template_path = Path(template_path) if template_path else None
        self._ndpbm_rates: Dict[str, float] = {}  # currency → rate

    def export(
        self,
        entities: ShipmentEntities,
        shipment_id: str,
        output_path: Optional[str | Path] = None,
    ) -> Path:
        # Export ke Excel CEISA 4.0.
        if output_path is None:
            output_path = Path("output") / f"{shipment_id}_ceisa.xlsx"
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Build all CEISA sheets
        for sheet_name in EXCEL_SHEET_ORDER:
            ws = wb.create_sheet(title=sheet_name)
            columns = EXCEL_COLUMNS.get(sheet_name, [])

            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")

            self._fill_sheet(ws, entities, shipment_id, sheet_name, columns)

        wb.save(output_path)
        logger.info(f"CEISA Excel exported: {output_path}")
        return output_path

    # SHEET FILLERS

    def _fill_sheet(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        sheet_name: str,
        columns: List[str],
    ) -> None:
        # Route ke sheet filler yang sesuai.
        fillers = {
            "HEADER":            self._fill_header,
            "ENTITAS":           self._fill_entitas,
            "DOKUMEN":           self._fill_dokumen,
            "PENGANGKUT":        self._fill_pengangkut,
            "KEMASAN":           self._fill_kemasan,
            "KONTAINER":         self._fill_kontainer,
            "KOMPONENBIAYA":     self._fill_komponenbiaya,
            "BARANG":            self._fill_barang,
            "PUNGUTAN":          self._fill_pungutan,
            "VERSI":             self._fill_versi,
        }
        filler = fillers.get(sheet_name)
        if filler:
            filler(ws, entities, shipment_id, columns)

    def _col(self, name: str, columns: List[str]) -> Optional[int]:
        # Get index kolom berdasarkan nama.
        try:
            return columns.index(name) + 1
        except ValueError:
            return None

    def _set(
        self,
        ws,
        row: int,
        col_name: str,
        value: Any,
        columns: List[str],
    ) -> None:
        # Tulis nilai ke (row, col_name).
        idx = self._col(col_name, columns)
        if idx is not None:
            ws.cell(row=row, column=idx, value=value)

    # HEADER

    def _fill_header(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        r = 2  # Data row

        # Identity dokumen
        self._set(ws, r, "NOMOR AJU", shipment_id, columns)
        self._set(ws, r, "KODE DOKUMEN", "20", columns)         # PIB declaration
        self._set(ws, r, "KODE KANTOR", DEFAULT_IMPORTER["kode_kantor"], columns)
        self._set(ws, r, "KODE JENIS IMPOR", "1", columns)     # Biasa
        self._set(ws, r, "KODE JENIS PROSEDUR", "1", columns)  # Biasa
        self._set(ws, r, "KODE CARA BAYAR", DEFAULT_IMPORTER["kode_cara_bayar"], columns)
        self._set(ws, r, "KODE TUTUP PU", "11", columns)

        # Mata uang dan incoterm
        currency = lk.get_currency_code(entities.currency or "USD")
        self._set(ws, r, "KODE VALUTA", currency, columns)

        incoterm = lk.get_incoterm_code(entities.incoterms)
        if incoterm:
            self._set(ws, r, "KODE INCOTERM", incoterm, columns)

        # Get NDPBM rate
        ndpbm = self._get_ndpbm(currency)

        # Financial calculations
        total_raw = br.parse_amount(entities.total_amount or "")
        freight_raw = br.parse_amount(entities.freight or "")
        invoice_currency = lk.get_currency_code(entities.currency or "USD")
        cif_based = lk.is_cif_based(incoterm or "")

        # NOTE: Indonesian customs practice for CIP/CIF imports — when freight is not
        # explicitly declared in the documents (freight_raw=0), FOB in the HEADER is
        # set to 0. The total CIF is declared as-is. When freight IS explicitly
        # extracted (even if 0 from a document field), calculate FOB properly.
        if cif_based:
            cif_usd = total_raw
            insurance_usd = br.calculate_insurance(cif_usd)
            if freight_raw > 0:
                # Freight explicitly declared
                fob_usd = max(0.0, cif_usd - freight_raw - insurance_usd)
            else:
                # No freight in documents — use FOB=0 per Indonesian customs practice
                fob_usd = 0.0
                freight_raw = 0.0
        else:
            # For FOB/EXW/FCA incoterms: total_amount IS the FOB price.
            fob_usd = total_raw
            insurance_usd = br.calculate_insurance(fob_usd)
            cif_usd = br.calculate_cif(fob_usd, freight_raw, insurance_usd)

        # Write financial fields
        self._set(ws, r, "FOB", round(fob_usd, 2), columns)
        self._set(ws, r, "CIF", round(cif_usd, 2), columns)
        self._set(ws, r, "FREIGHT", round(freight_raw, 2), columns)
        self._set(ws, r, "NDPBM", ndpbm, columns)

        # Insurance code: 'DN' for domestic, 'LN' for overseas
        if cif_based and freight_raw == 0.0:
            insurance_usd = 0.0
        self._set(ws, r, "KODE ASURANSI", "DN", columns)
        self._set(ws, r, "ASURANSI", round(insurance_usd, 2), columns)

        # Ports
        port_load = lk.get_port_locode(entities.port_of_loading or "")
        port_discharge = lk.get_port_locode(entities.port_of_discharge or "")

        if port_load:
            self._set(ws, r, "KODE PELABUHAN MUAT", port_load, columns)
            self._set(ws, r, "KODE PELABUHAN TRANSIT", port_load, columns)
            self._set(ws, r, "KODE PELABUHAN MUAT AKHIR", port_load, columns)

        if port_discharge:
            self._set(ws, r, "KODE PELABUHAN BONGKAR", port_discharge, columns)
            self._set(ws, r, "KODE PELABUHAN TUJUAN", port_discharge, columns)

        # Weights
        if entities.total_gross_weight:
            gw = br.parse_amount(entities.total_gross_weight)
            self._set(ws, r, "BRUTO", round(gw, 3), columns)

        if entities.total_net_weight:
            nw = br.parse_amount(entities.total_net_weight)
            self._set(ws, r, "NETTO", round(nw, 3), columns)
        elif entities.items:
            total_nw = sum(br.parse_amount(i.net_weight or "0") for i in entities.items)
            if total_nw > 0:
                self._set(ws, r, "NETTO", round(total_nw, 3), columns)

        # Declarant / PJTP
        self._set(ws, r, "KOTA PERNYATAAN", "CIKARANG", columns)
        self._set(ws, r, "NAMA PERNYATAAN", DEFAULT_IMPORTER["name"], columns)
        self._set(ws, r, "JABATAN PERNYATAAN", "MANAGER IMPORT", columns)
        self._set(ws, r, "KODE GUNA BARANG", "KMD", columns)
        self._set(ws, r, "KODE ASAL BARANG", "1", columns)   # 1 = local content
        self._set(ws, r, "FLAG PROPORSIONAL NETTO", "T", columns)

    # ENTITAS

    def _fill_entitas(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        r = 2

        def write(
            seri: int,
            kode: int,
            nama: Optional[str],
            alamat: Optional[str] = "",
            negara: str = "CN",
            npwp: Optional[str] = None,
            nib: Optional[str] = None,
            api: Optional[str] = None,
        ):
            nonlocal r
            if not nama:
                return
            self._set(ws, r, "NOMOR AJU", shipment_id, columns)
            self._set(ws, r, "SERI", str(seri), columns)
            self._set(ws, r, "KODE ENTITAS", str(kode), columns)
            self._set(ws, r, "NAMA ENTITAS", nama.strip(), columns)
            if alamat:
                self._set(ws, r, "ALAMAT ENTITAS", alamat.strip(), columns)
            if npwp:
                self._set(ws, r, "NOMOR IDENTITAS", npwp, columns)
            if nib:
                self._set(ws, r, "NIB ENTITAS", nib, columns)
            if api:
                self._set(ws, r, "KODE JENIS API", api, columns)

            country_code = lk.get_country_code(negara)
            self._set(ws, r, "KODE NEGARA", country_code, columns)
            r += 1

        # Seller / Eksportir (Seri=9, Kode=9)
        seller_country = entities.country_of_origin or "CN"
        if entities.seller_name:
            write(
                seri=9, kode=9,
                nama=entities.seller_name,
                alamat=entities.seller_address,
                negara=seller_country,
            )
            # Manufacturer (Seri=10, Kode=10)
            write(
                seri=10, kode=10,
                nama=entities.seller_name,
                alamat=entities.seller_address,
                negara=seller_country,
            )

        # Buyer / Importir (Seri=1, Kode=1)
        if entities.buyer_name:
            write(
                seri=1, kode=1,
                nama=entities.buyer_name,
                alamat=entities.buyer_address,
                negara="ID",
                npwp=DEFAULT_IMPORTER["npwp"],
                nib=DEFAULT_IMPORTER["nib"],
                api=DEFAULT_IMPORTER["api_type"],
            )

        # Shipper (Seri=7, Kode=7), only write if shipper differs from seller
        shipper_country = entities.country_of_origin or "CN"
        if entities.shipper_name and entities.shipper_name != entities.seller_name:
            write(
                seri=7, kode=7,
                nama=entities.shipper_name,
                alamat=entities.shipper_address,
                negara=shipper_country,
            )

        # Notify Party (Seri=4, Kode=4)
        notify_name = entities.notify_party_name or entities.consignee_name
        if notify_name:
            write(
                seri=4, kode=4,
                nama=notify_name,
                alamat=entities.notify_party_address or entities.consignee_address,
                negara="ID",
            )

        # Consignee (Seri=11, Kode=11)
        if entities.consignee_name and entities.consignee_name != entities.buyer_name:
            write(
                seri=11, kode=11,
                nama=entities.consignee_name,
                alamat=entities.consignee_address,
                negara="ID",
            )

    # DOKUMEN

    def _fill_dokumen(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        r = 2

        for seri, kode, nomor, tanggal in [
            (1, "380", entities.invoice_number, entities.invoice_date),
            (2, "860", entities.bl_number, entities.bl_date),
        ]:
            if not nomor:
                continue
            parsed_date = br.parse_date(tanggal or "")
            self._set(ws, r, "NOMOR AJU", shipment_id, columns)
            self._set(ws, r, "SERI", f"{seri}.0", columns)
            self._set(ws, r, "KODE DOKUMEN", kode, columns)
            self._set(ws, r, "NOMOR DOKUMEN", str(nomor).strip(), columns)
            if parsed_date:
                self._set(ws, r, "TANGGAL DOKUMEN", parsed_date, columns)
            r += 1

    # PENGANGKUT

    def _fill_pengangkut(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        if not (entities.vessel_name or entities.voyage_number):
            return
        r = 2
        self._set(ws, r, "NOMOR AJU", shipment_id, columns)
        self._set(ws, r, "SERI", "1", columns)
        self._set(ws, r, "KODE CARA ANGKUT", "1", columns)
        if entities.vessel_name:
            self._set(ws, r, "NAMA PENGANGKUT", entities.vessel_name.strip(), columns)
        if entities.voyage_number:
            self._set(ws, r, "NOMOR PENGANGKUT", entities.voyage_number.strip(), columns)

    # KEMASAN

    def _fill_kemasan(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        # Use item-level packaging if available
        pkg_type = lk.get_packaging_code(entities.packaging_type or "")
        nop = br.parse_amount(entities.number_of_packages or "")

        # If no extracted packages, sum from items
        if nop == 0 and entities.items:
            total_qty = sum(br.parse_amount(str(i.quantity or "0")) for i in entities.items)
            if total_qty > 0:
                nop = total_qty
                pkg_type = "PCE"

        if nop == 0:
            return

        r = 2
        self._set(ws, r, "NOMOR AJU", shipment_id, columns)
        self._set(ws, r, "SERI", "1", columns)
        self._set(ws, r, "KODE KEMASAN", pkg_type, columns)
        self._set(ws, r, "JUMLAH KEMASAN", nop, columns)
        self._set(ws, r, "MEREK KEMASAN", "-", columns)

    # KONTAINER

    def _fill_kontainer(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        # Fill KONTAINER (container) sheet.
        if not entities.container_numbers:
            return

        r = 2
        for i, container in enumerate(entities.container_numbers, start=1):
            # Parse container number for type hints
            parsed = br.parse_container_number(container)
            size_code = br.parse_container_size(container)

            self._set(ws, r, "NOMOR AJU", shipment_id, columns)
            self._set(ws, r, "SERI", str(i), columns)
            self._set(ws, r, "NOMOR KONTINER", container.upper().strip(), columns)
            self._set(ws, r, "KODE UKURAN KONTAINER", size_code, columns)
            self._set(ws, r, "KODE JENIS KONTAINER", "8", columns)  # General
            self._set(ws, r, "KODE TIPE KONTAINER", "1", columns)  # Standard
            r += 1

        # Seal numbers (attach to each container)
        for i, seal in enumerate(entities.seal_numbers):
            target_row = 2 + i
            if target_row < r:
                self._set(ws, target_row, "NOMOR SEGEL", seal.upper().strip(), columns)

    # KOMPONENBIAYA

    def _fill_komponenbiaya(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        # Fill KOMPONENBIAYA (cost component) sheet.
        r = 2

        currency = lk.get_currency_code(entities.currency or "USD")
        ndpbm = self._get_ndpbm(currency)
        incoterm = lk.get_incoterm_code(entities.incoterms or "")

        total_raw = br.parse_amount(entities.total_amount or "")
        freight_raw = br.parse_amount(entities.freight or "")
        cif_based = lk.is_cif_based(incoterm)

        if cif_based:
            fob_usd = br.calc_fob_from_total(
                total_raw, currency, incoterm, freight_raw
            )
            if freight_raw == 0.0 and entities.container_numbers:
                freight_raw = br.get_freight_cost("40", len(entities.container_numbers))
        else:
            fob_usd = total_raw

        cif_usd = br.calculate_cif(fob_usd, freight_raw, 0.0)
        insurance_usd = br.calculate_insurance(cif_usd)
        cif_usd = br.calculate_cif(fob_usd, freight_raw, insurance_usd)

        # Row 1: Harga Invoice (Type=1)
        self._set(ws, r, "NOMOR AJU", shipment_id, columns)
        self._set(ws, r, "JENIS NILAI", "1", columns)
        self._set(ws, r, "HARGA INVOICE", round(fob_usd, 2), columns)
        self._set(ws, r, "BIAYA TRANSPORTASI", round(freight_raw, 2), columns)
        self._set(ws, r, "BIAYA ASURANSI", round(insurance_usd, 2), columns)

        cif_rupiah = cif_usd * ndpbm
        self._set(ws, r, "CIF", round(cif_usd, 2), columns)
        self._set(ws, r, "CIF RUPIAH", round(cif_rupiah, 2), columns)
        self._set(ws, r, "NDPBM", ndpbm, columns)

    # BARANG

    def _fill_barang(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        # Fill BARANG (items) sheet with full line-item data.
        currency = lk.get_currency_code(entities.currency or "USD")
        ndpbm = self._get_ndpbm(currency)
        negara_asal = lk.get_country_code(entities.country_of_origin or "CN")
        incoterm = lk.get_incoterm_code(entities.incoterms or "")

        total_fob_raw = br.parse_amount(entities.total_amount or "")
        freight_raw    = br.parse_amount(entities.freight or "")
        if freight_raw == 0.0 and entities.container_numbers:
            freight_raw = br.get_freight_cost("40", len(entities.container_numbers))

        r = 2
        for i, item in enumerate(entities.items, start=1):
            self._set(ws, r, "NOMOR AJU", shipment_id, columns)
            self._set(ws, r, "SERI BARANG", str(i), columns)

            # HS code
            hs = self._normalize_hs(item.hs_code)
            self._set(ws, r, "HS", hs, columns)

            # Item code
            if item.item_code:
                self._set(ws, r, "KODE BARANG", item.item_code.strip(), columns)

            # Description
            desc = (item.description or "-").strip()
            self._set(ws, r, "URAIAN", desc, columns)

            # Brand / model
            self._set(ws, r, "MEREK", item.brand or "N/A", columns)
            self._set(ws, r, "TIPE", item.model or "N/A", columns)

            # Dimensions
            if item.dimensions:
                self._set(ws, r, "UKURAN", item.dimensions.strip(), columns)

            self._set(ws, r, "SPESIFIKASI LAIN", "-", columns)

            # Unit of measure
            unit_code = lk.get_packaging_code(item.unit or "PCE")
            self._set(ws, r, "KODE SATUAN", unit_code, columns)

            # Quantity
            qty = br.parse_amount(item.quantity or "0")
            if qty > 0:
                self._set(ws, r, "JUMLAH SATUAN", qty, columns)

            # Packaging
            pkg_code = lk.get_packaging_code(item.packaging or "CT")
            self._set(ws, r, "KODE KEMASAN", pkg_code, columns)

            cartons = br.parse_amount(item.cartons or "0")
            if cartons > 0:
                self._set(ws, r, "JUMLAH KEMASAN", cartons, columns)

            # Weights
            nw = br.parse_amount(item.net_weight or "0")
            if nw > 0:
                self._set(ws, r, "NETTO", round(nw, 3), columns)

            gw = br.parse_amount(item.gross_weight or "0")
            if gw > 0:
                self._set(ws, r, "BRUTO", round(gw, 3), columns)

            # Financial values
            fob_item = br.parse_amount(item.amount or "0")

            if fob_item == 0:
                q = br.parse_amount(item.quantity or "0")
                p = br.parse_amount(item.unit_price or "0")
                if q > 0 and p > 0:
                    fob_item = q * p

            if fob_item > 0:
                self._set(ws, r, "FOB", round(fob_item, 2), columns)
                self._set(ws, r, "CIF", round(fob_item, 2), columns)
                self._set(ws, r, "CIF RUPIAH", round(fob_item * ndpbm, 2), columns)
                self._set(ws, r, "NDPBM", ndpbm, columns)

            # Unit price
            up = br.parse_amount(item.unit_price or "0")
            if up > 0:
                self._set(ws, r, "HARGA SATUAN", round(up, 4), columns)

            # Fixed fields
            self._set(ws, r, "KODE GUNA BARANG", "KMD", columns)
            self._set(ws, r, "KODE ASAL BARANG", "1", columns)
            self._set(ws, r, "KODE NEGARA ASAL", negara_asal, columns)
            self._set(ws, r, "PERNYATAAN LARTAS", "T", columns)
            self._set(ws, r, "FLAG 4 TAHUN", "T", columns)
            self._set(ws, r, "KODE KONDISI BARANG", "1", columns)
            self._set(ws, r, "METODE PENENTUAN NILAI", "Metode 1", columns)
            self._set(ws, r, "STATEMENT PERBEDAAN HARGA", "T", columns)

            r += 1

    # PUNGUTAN

    def _fill_pungutan(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        currency = lk.get_currency_code(entities.currency or "USD")
        ndpbm = self._get_ndpbm(currency)
        incoterm = lk.get_incoterm_code(entities.incoterms or "")

        total_raw = br.parse_amount(entities.total_amount or "")
        freight_raw = br.parse_amount(entities.freight or "")
        cif_based = lk.is_cif_based(incoterm)

        if cif_based:
            fob_usd = br.calc_fob_from_total(
                total_raw, currency, incoterm, freight_raw
            )
            if freight_raw == 0.0 and entities.container_numbers:
                freight_raw = br.get_freight_cost("40", len(entities.container_numbers))
        else:
            fob_usd = total_raw

        # Insurance
        cif_usd_before = br.calculate_cif(fob_usd, freight_raw, 0.0)
        insurance_usd = br.calculate_insurance(cif_usd_before)
        cif_usd = br.calculate_cif(fob_usd, freight_raw, insurance_usd)

        # Determine BM rate from first item's HS code
        bm_tariff_pct = 5.0  # default
        if entities.items:
            first_hs = entities.items[0].hs_code
            bm_tariff_pct, _ = lk.get_hs_tariff(first_hs)

        bm_usd = br.calculate_bm(fob_usd, freight_raw, insurance_usd, bm_tariff_pct)
        ppn_usd = br.calculate_ppn(bm_usd, cif_usd)

        r = 2

        # Row 1: BM (Bea Masuk)
        self._set(ws, r, "NOMOR AJU", shipment_id, columns)
        self._set(ws, r, "KODE FASILITAS TARIF", "1", columns)
        self._set(ws, r, "KODE JENIS PUNGUTAN", "BM", columns)
        self._set(ws, r, "JUMLAH PUNGUTAN", round(bm_usd, 2), columns)
        self._set(ws, r, "KODE TARIF", f"BM{bm_tariff_pct:.0f}%", columns)
        r += 1

        # Row 2: PPN
        self._set(ws, r, "NOMOR AJU", shipment_id, columns)
        self._set(ws, r, "KODE FASILITAS TARIF", "1", columns)
        self._set(ws, r, "KODE JENIS PUNGUTAN", "PPN", columns)
        self._set(ws, r, "JUMLAH PUNGUTAN", round(ppn_usd, 2), columns)
        r += 1

        # Row 3: PPNBM (if applicable — 0 for now)
        # self._set(ws, r, "NOMOR AJU", shipment_id, columns)
        # self._set(ws, r, "KODE JENIS PUNGUTAN", "PPNBM", columns)
        # self._set(ws, r, "JUMLAH PUNGUTAN", 0.0, columns)

    # VERSI

    def _fill_versi(
        self,
        ws,
        entities: ShipmentEntities,
        shipment_id: str,
        columns: List[str],
    ) -> None:
        self._set(ws, 2, "VERSION", "1.3", columns)

    # Helpers

    def _get_ndpbm(self, currency: str) -> float:
        if currency in self._ndpbm_rates:
            return self._ndpbm_rates[currency]
        rate = br.get_ndpbm_rate(currency)
        self._ndpbm_rates[currency] = rate
        return rate

    def _normalize_hs(self, hs_code: Optional[str]) -> str:
        if not hs_code:
            return "94031000"
        cleaned = re.sub(r"[^0-9]", "", str(hs_code))
        if len(cleaned) >= 6:
            return cleaned[:8].ljust(8, "0")
        return "94031000"
