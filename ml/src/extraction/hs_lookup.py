from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Dict, Optional, Tuple

logger = logging.getLogger(__name__)

_cache: Dict[int, Tuple[float, float, float]] = {}
_cache_loaded = False


def _default_rates() -> Dict[str, float]:
    return {"bm": 0.0, "ppn": 11.0, "pph": 0.0, "hs_code": None, "found": False}


def _ensure_master_loaded() -> None:
    global _cache_loaded
    if _cache_loaded:
        return
    _load_master()
    _cache_loaded = True


def _load_master() -> None:
    # Load MASTER CEK HS CODE.xlsx into the module cache.
    try:
        import openpyxl

        search_paths = [
            Path(__file__).parent.parent.parent / "MASTER CEK HS CODE.xlsx",
            Path(__file__).parent.parent.parent.parent / "MASTER CEK HS CODE.xlsx",
            Path.cwd() / "MASTER CEK HS CODE.xlsx",
        ]

        path = None
        for p in search_paths:
            if p.exists():
                path = p
                break

        if path is None:
            logger.warning(
                f"MASTER CEK HS CODE.xlsx not found in {search_paths}. "
                "HS lookup will return zeros."
            )
            return

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        sheet_name = (
            "DATA MASTER HS CODE"
            if "DATA MASTER HS CODE" in wb.sheetnames
            else "CEK HS"
        )
        ws = wb[sheet_name]

        entries = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                # col[3] = HS code, col[6] = BM, col[7] = PPN, col[8] = PPH
                raw_hs = row[3]
                if raw_hs is None:
                    continue
                code_str = str(raw_hs).strip()
                code_int = int(code_str[:8].ljust(8, "0"))

                bm_raw = row[6] if len(row) > 6 else None
                ppn_raw = row[7] if len(row) > 7 else None
                pph_raw = row[8] if len(row) > 8 else None

                bm = float(bm_raw) * 100 if bm_raw is not None else 0.0
                ppn = float(ppn_raw) * 100 if ppn_raw is not None else 0.0
                pph = float(pph_raw) * 100 if pph_raw is not None else 0.0

                _cache[code_int] = (
                    round(bm, 4),
                    round(ppn, 4),
                    round(pph, 4),
                )
                entries += 1
            except (ValueError, TypeError, IndexError):
                continue

        wb.close()
        logger.info(f"HS master loaded: {entries} entries from '{sheet_name}' sheet of {path.name}")
    except Exception as e:
        logger.warning(f"Failed to load HS master: {e}. HS lookup will return zeros.")


def get_hs_tariff(hs_code: int | str) -> Dict[str, float]:
    _ensure_master_loaded()

    try:
        if isinstance(hs_code, str):
            code_int = int(str(hs_code).strip()[:8].ljust(8, "0"))
        else:
            code_int = int(hs_code)
    except (ValueError, TypeError):
        return _default_rates()

    # Exact 8-digit match
    if code_int in _cache:
        bm, ppn, pph = _cache[code_int]
        return {
            "bm": bm, "ppn": ppn, "pph": pph,
            "hs_code": str(code_int).zfill(8),
            "found": True,
        }

    # 4-digit prefix match
    prefix4 = code_int // 10000
    for cached_code, (bm, ppn, pph) in _cache.items():
        if cached_code // 10000 == prefix4:
            logger.info(f"HS {code_int}: using 4-digit prefix match {cached_code}")
            return {
                "bm": bm, "ppn": ppn, "pph": pph,
                "hs_code": str(code_int).zfill(8),
                "found": False,
            }

    # Default rates
    return _default_rates()


def clear_cache() -> None:
    global _cache_loaded
    _cache.clear()
    _cache_loaded = False
