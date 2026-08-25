from __future__ import annotations

import json
import logging
import re
import string
from collections import defaultdict
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import fitz
import io
import openpyxl

logger = logging.getLogger(__name__)

# BIO Label Scheme

LABEL_TO_ID: Dict[str, int] = {
    "O": 0,
    # Document identifiers
    "B-invoice_number": 1, "I-invoice_number": 2,
    "B-invoice_date": 3,   "I-invoice_date": 4,
    "B-bl_number": 5,       "I-bl_number": 6,
    "B-bl_date": 7,         "I-bl_date": 8,
    # Parties
    "B-seller_name": 9,     "I-seller_name": 10,
    "B-seller_address": 11, "I-seller_address": 12,
    "B-buyer_name": 13,     "I-buyer_name": 14,
    "B-buyer_address": 15,  "I-buyer_address": 16,
    "B-shipper_name": 17,   "I-shipper_name": 18,
    "B-shipper_address": 19,"I-shipper_address": 20,
    "B-consignee_name": 21, "I-consignee_name": 22,
    "B-consignee_address": 23,"I-consignee_address": 24,
    "B-notify_party_name": 25, "I-notify_party_name": 26,
    "B-notify_party_address": 27,"I-notify_party_address": 28,
    # Transportation
    "B-vessel_name": 29,    "I-vessel_name": 30,
    "B-voyage_number": 31,  "I-voyage_number": 32,
    "B-port_of_loading": 33,"I-port_of_loading": 34,
    "B-port_of_discharge": 35,"I-port_of_discharge": 36,
    "B-place_of_receipt": 37,"I-place_of_receipt": 38,
    "B-place_of_delivery": 39,"I-place_of_delivery": 40,
    # Financial
    "B-currency": 41,       "I-currency": 42,
    "B-incoterms": 43,      "I-incoterms": 44,
    "B-freight_term": 45,   "I-freight_term": 46,
    "B-country_of_origin": 47,"I-country_of_origin": 48,
    "B-country_of_destination": 49,"I-country_of_destination": 50,
    # Summary totals
    "B-total_amount": 51,    "I-total_amount": 52,
    "B-total_quantity": 53,  "I-total_quantity": 54,
    "B-total_net_weight": 55,"I-total_net_weight": 56,
    "B-total_gross_weight": 57,"I-total_gross_weight": 58,
    "B-number_of_packages": 59,"I-number_of_packages": 60,
    "B-cbm": 61,            "I-cbm": 62,
    # Line item entities (Pattern layer, but useful for LayoutXLM too)
    "B-item_description": 63,"I-item_description": 64,
    "B-item_hs_code": 65,   "I-item_hs_code": 66,
    "B-item_quantity": 67,  "I-item_quantity": 68,
    "B-item_unit": 69,      "I-item_unit": 70,
    "B-item_unit_price": 71,"I-item_unit_price": 72,
    "B-item_amount": 73,    "I-item_amount": 74,
    "B-item_net_weight": 75,"I-item_net_weight": 76,
    "B-container_number": 77,"I-container_number": 78,
    "B-seal_number": 79,    "I-seal_number": 80,
}

ID_TO_LABEL: Dict[int, str] = {v: k for k, v in LABEL_TO_ID.items()}
NUM_LABELS = len(LABEL_TO_ID)

# Maps canonical entity name → (Excel sheet, column name)
ENTITY_EXCEL_MAP: Dict[str, Tuple[str, str]] = {
    # HEADER
    "invoice_number": ("HEADER", "NOMOR DOKUMEN"),
    "invoice_date": ("HEADER", "TANGGAL DOKUMEN"),
    "currency": ("HEADER", "KODE VALUTA"),
    "incoterms": ("HEADER", "KODE INCOTERM"),
    "port_of_loading": ("HEADER", "KODE PELABUHAN MUAT"),
    "port_of_discharge": ("HEADER", "KODE PELABUHAN BONGKAR"),
    "total_amount": ("HEADER", "FOB"),
    "freight": ("HEADER", "FREIGHT"),
    "total_gross_weight": ("HEADER", "BRUTO"),
    "total_net_weight": ("HEADER", "NETTO"),
    # ENTITAS
    "seller_name": ("ENTITAS", "NAMA ENTITAS"),        # Kode=9
    "seller_address": ("ENTITAS", "ALAMAT ENTITAS"),   # Kode=9
    "buyer_name": ("ENTITAS", "NAMA ENTITAS"),         # Kode=1
    "buyer_address": ("ENTITAS", "ALAMAT ENTITAS"),    # Kode=1
    "shipper_name": ("ENTITAS", "NAMA ENTITAS"),       # Kode=7
    "shipper_address": ("ENTITAS", "ALAMAT ENTITAS"),  # Kode=7
    "consignee_name": ("ENTITAS", "NAMA ENTITAS"),     # Kode=11
    "consignee_address": ("ENTITAS", "ALAMAT ENTITAS"),# Kode=11
    "notify_party_name": ("ENTITAS", "NAMA ENTITAS"),  # Kode=4
    "notify_party_address": ("ENTITAS", "ALAMAT ENTITAS"), # Kode=4
    "vessel_name": ("PENGANGKUT", "NAMA PENGANGKUT"),
    "voyage_number": ("PENGANGKUT", "NOMOR PENGANGKUT"),
    # DOKUMEN
    "bl_number": ("DOKUMEN", "NOMOR DOKUMEN"),
    "bl_date": ("DOKUMEN", "TANGGAL DOKUMEN"),
    # KEMASAN
    "number_of_packages": ("KEMASAN", "JUMLAH KEMASAN"),
    "packaging_type": ("KEMASAN", "KODE KEMASAN"),
    # KONTAINER
    "container_numbers": ("KONTAINER", "NOMOR KONTINER"),
    "seal_numbers": ("KONTAINER", "NOMOR SEGEL"),
    # BARANG (first item only for summary; full table handled separately)
    "item_description": ("BARANG", "URAIAN"),
    "item_hs_code": ("BARANG", "HS"),
    "item_quantity": ("BARANG", "JUMLAH SATUAN"),
    "item_unit": ("BARANG", "KODE SATUAN"),
    "item_unit_price": ("BARANG", "HARGA SATUAN"),
    "item_amount": ("BARANG", "FOB"),
    "item_net_weight": ("BARANG", "NETTO"),
    "item_gross_weight": ("BARANG", "BRUTO"),
}

# Kode ENTITAS mapping (sheet=ENTITAS, kode column value)
ENTITAS_KODE_MAP: Dict[str, int] = {
    "seller_name": 9,
    "seller_address": 9,
    "buyer_name": 1,
    "buyer_address": 1,
    "shipper_name": 7,
    "shipper_address": 7,
    "consignee_name": 11,
    "consignee_address": 11,
    "notify_party_name": 4,
    "notify_party_address": 4,
}

# DOKUMEN kode (sheet=DOKUMEN, kode_dokumen column value)
DOKUMEN_KODE_MAP: Dict[str, str] = {
    "invoice_number": "380",
    "invoice_date": "380",
    "bl_number": "860",
    "bl_date": "860",
}


@dataclass
class OCRWord:
    text: str
    x0: float; y0: float; x1: float; y1: float
    block_num: int = 0
    line_num: int = 0
    word_num: int = 0

    @property
    def bbox(self) -> List[float]:
        return [self.x0, self.y0, self.x1, self.y1]

    def cleaned_text(self) -> str:
        return self.text.strip()


@dataclass
class OCRPage:
    page_num: int
    width: float
    height: float
    words: List[OCRWord]

    @property
    def text_lines(self) -> List[str]:
        # Reconstruct text lines in reading order.
        if not self.words:
            return []
        lines: List[List[OCRWord]] = []
        current_line: List[OCRWord] = []
        last_y = -1
        for w in self.words:
            y_key = round(w.y0, 1)
            if last_y < 0 or abs(y_key - last_y) < 5:
                current_line.append(w)
                last_y = y_key
            else:
                if current_line:
                    lines.append(current_line)
                current_line = [w]
                last_y = y_key
        if current_line:
            lines.append(current_line)
        return [" ".join(w.cleaned_text() for w in line) for line in lines]


@dataclass
class EntitySpan:
    # A matched entity span in OCR text.
    entity_name: str
    value: str
    start_word_idx: int
    end_word_idx: int  # inclusive
    matched_text: str
    similarity: float


@dataclass
class LabeledPage:
    # A training sample: page with per-word labels.
    shipment_id: str
    doc_type: str  # "CI", "PL", "BL"
    page_num: int
    width: float
    height: float
    words: List[OCRWord]
    labels: List[int]
    spans: List[EntitySpan]


class GroundTruthReader:
    def __init__(self, excel_path: str | Path):
        self.excel_path = Path(excel_path)
        self._cache: Dict[str, Any] = {}
        self._load()

    def _load(self) -> None:
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h).strip() if h else "" for h in rows[0]]
            data = [dict(zip(headers, row)) for row in rows[1:] if any(row)]
            self._cache[sheet_name] = data
        wb.close()
        logger.debug(f"Loaded {len(self._cache)} sheets from {self.excel_path.name}")

    def get_header_value(self, column: str) -> Optional[str]:
        rows = self._cache.get("HEADER", [])
        for row in rows:
            val = row.get(column)
            if val is not None and str(val).strip():
                return str(val).strip()
        return None

    def get_entitas(self, kode: int) -> Optional[Dict[str, str]]:
        rows = self._cache.get("ENTITAS", [])
        for row in rows:
            kode_val = row.get("KODE ENTITAS")
            try:
                if int(str(kode_val or 0)) == kode:
                    return {k: str(v).strip() if v else "" for k, v in row.items()}
            except (ValueError, TypeError):
                pass
        return None

    def get_dokumen(self, kode: str) -> Optional[Dict[str, str]]:
        rows = self._cache.get("DOKUMEN", [])
        for row in rows:
            kode_val = row.get("KODE DOKUMEN")
            if str(kode_val or "") == str(kode):
                return {k: str(v).strip() if v else "" for k, v in row.items()}
        return None

    def get_pengangkut(self) -> Optional[Dict[str, str]]:
        rows = self._cache.get("PENGANGKUT", [])
        if rows:
            return {k: str(v).strip() if v else "" for k, v in rows[0].items()}
        return None

    def get_kemasan(self) -> Optional[Dict[str, str]]:
        rows = self._cache.get("KEMASAN", [])
        if rows:
            return {k: str(v).strip() if v else "" for k, v in rows[0].items()}
        return None

    def get_kontainer(self) -> List[Dict[str, str]]:
        rows = self._cache.get("KONTAINER", [])
        return [{k: str(v).strip() if v else "" for k, v in row.items()} for row in rows]

    def get_barang(self) -> List[Dict[str, str]]:
        rows = self._cache.get("BARANG", [])
        return [{k: str(v).strip() if v else "" for k, v in row.items()} for row in rows]

    def get_all_entities(self) -> Dict[str, List[str]]:
        entities: Dict[str, List[str]] = defaultdict(list)

        # ENTITAS entities (by kode)
        entitas_map = {
            9: "seller", 1: "buyer", 7: "shipper", 11: "consignee", 4: "notify_party"
        }
        for kode, prefix in entitas_map.items():
            row = self.get_entitas(kode)
            if row:
                name = row.get("NAMA ENTITAS", "")
                addr = row.get("ALAMAT ENTITAS", "")
                if name:
                    entities[f"{prefix}_name"].append(name)
                if addr:
                    entities[f"{prefix}_address"].append(addr)

        # DOKUMEN entities
        for dok_kode, entity in [("380", "invoice"), ("860", "bl")]:
            row = self.get_dokumen(dok_kode)
            if row:
                num = row.get("NOMOR DOKUMEN", "")
                date = row.get("TANGGAL DOKUMEN", "")
                if num:
                    entities[f"{entity}_number"].append(num)
                if date:
                    entities[f"{entity}_date"].append(str(date)[:10])

        # PENGANGKUT
        row = self.get_pengangkut()
        if row:
            vessel = row.get("NAMA PENGANGKUT", "")
            voyage = row.get("NOMOR PENGANGKUT", "")
            if vessel:
                entities["vessel_name"].append(vessel)
            if voyage:
                entities["voyage_number"].append(voyage)

        # KEMASAN
        row = self.get_kemasan()
        if row:
            nop = row.get("JUMLAH KEMASAN", "")
            pkg = row.get("KODE KEMASAN", "")
            if nop:
                entities["number_of_packages"].append(nop)
            if pkg:
                entities["packaging_type"].append(pkg)

        # KONTAINER
        for row in self.get_kontainer():
            cn = row.get("NOMOR KONTINER", "")
            seal = row.get("NOMOR SEGEL", "")
            if cn:
                entities["container_numbers"].append(cn)
            if seal:
                entities["seal_numbers"].append(seal)

        # BARANG (summary: first item HS, last item total)
        barang_rows = self.get_barang()
        for row in barang_rows:
            hs = row.get("HS", "")
            desc = row.get("URAIAN", "")
            qty = row.get("JUMLAH SATUAN", "")
            unit = row.get("KODE SATUAN", "")
            price = row.get("HARGA SATUAN", "")
            amount = row.get("FOB", "")
            nw = row.get("NETTO", "")
            gw = row.get("BRUTO", "")
            if hs:
                entities["item_hs_code"].append(hs)
            if desc:
                entities["item_description"].append(desc)
            if qty:
                entities["item_quantity"].append(qty)
            if unit:
                entities["item_unit"].append(unit)
            if price:
                entities["item_unit_price"].append(price)
            if amount:
                entities["item_amount"].append(amount)
            if nw:
                entities["item_net_weight"].append(nw)
            if gw:
                entities["item_gross_weight"].append(gw)

        # HEADER
        header_map = {
            "currency": "KODE VALUTA",
            "incoterms": "KODE INCOTERM",
            "total_amount": "FOB",
            "freight": "FREIGHT",
            "total_gross_weight": "BRUTO",
            "total_net_weight": "NETTO",
            "port_of_loading": "KODE PELABUHAN MUAT",
            "port_of_discharge": "KODE PELABUHAN BONGKAR",
            "country_of_origin": "KODE ASAL BARANG",
        }
        for entity_name, col_name in header_map.items():
            val = self.get_header_value(col_name)
            if val:
                entities[entity_name].append(val)

        return dict(entities)


class PDFOCRReader:
    def __init__(
        self,
        use_paddle_fallback: bool = True,
        ocr_results: Optional[Dict[str, Any]] = None,
    ):
        self.use_paddle_fallback = use_paddle_fallback
        self.ocr_results = ocr_results or {}
        self._paddle_cache: Dict[str, Any] = {}

    def read_pdf(self, pdf_path: str | Path) -> List[OCRPage]:
        pdf_path = Path(pdf_path)
        pages: List[OCRPage] = []

        doc = fitz.open(str(pdf_path))
        for page_num in range(len(doc)):
            page = doc[page_num]
            width = page.rect.width
            height = page.rect.height

            # Try pre-computed OCR results first
            img_key = f"{pdf_path.stem}_p{page_num}.png"
            found_in_precomputed = False
            for img_path_str, ocr_data in self.ocr_results.items():
                if img_path_str.endswith(img_key):
                    words_data = ocr_data.get("words", [])
                    words = [
                        OCRWord(
                            text=w["text"],
                            x0=w["bbox"][0], y0=w["bbox"][1],
                            x1=w["bbox"][2], y1=w["bbox"][3],
                        )
                        for w in words_data
                    ]
                    pages.append(OCRPage(
                        page_num=page_num,
                        width=width, height=height,
                        words=words,
                    ))
                    found_in_precomputed = True
                    break

            if found_in_precomputed:
                continue

            # Live OCR: PyMuPDF first (fast for text-based PDFs)
            raw_words = page.get_text("words")
            if raw_words:
                words = [
                    OCRWord(
                        text=w[4],
                        x0=w[0], y0=w[1], x1=w[2], y1=w[3],
                        block_num=w[5] if len(w) > 5 else 0,
                        line_num=w[6] if len(w) > 6 else 0,
                        word_num=w[7] if len(w) > 7 else 0,
                    )
                    for w in raw_words
                ]
            elif self.use_paddle_fallback:
                logger.info(f"  PaddleOCR fallback for page {page_num} ({pdf_path.name})")
                words = self._paddle_ocr_page(pdf_path, page_num, width, height)
            else:
                words = []

            pages.append(OCRPage(
                page_num=page_num,
                width=width, height=height,
                words=words,
            ))

        doc.close()
        return pages

    def _paddle_ocr_page(
        self,
        pdf_path: str | Path,
        page_num: int,
        width: float,
        height: float,
        timeout: int = 60,
    ) -> List[OCRWord]:
        import numpy as np
        import PIL.Image

        try:
            from paddleocr import PaddleOCR
        except ImportError:
            logger.warning("PaddleOCR not installed")
            return []

        try:
            cache_key = str(pdf_path)
            if cache_key not in self._paddle_cache:
                logger.info(f"  Initializing PaddleOCR for {Path(pdf_path).name} (first use)")
                self._paddle_cache[cache_key] = PaddleOCR(
                    lang="en", use_angle_cls=False, show_log=False, use_gpu=False
                )
            ocr = self._paddle_cache[cache_key]

            # Render PDF page to image using PyMuPDF
            doc = fitz.open(str(pdf_path))
            page = doc[page_num]
            mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better OCR
            pix = page.get_pixmap(matrix=mat)
            img_bytes = pix.tobytes("png")
            doc.close()

            # Convert to numpy array
            img = PIL.Image.open(io.BytesIO(img_bytes))
            img_array = np.array(img)

            # Run PaddleOCR (CPU mode)
            result = ocr.ocr(img_array, cls=False)
            words = []
            if result and result[0]:
                for line in result[0]:
                    box = line[0]
                    text = line[1][0]
                    x0 = min(p[0] for p in box)
                    y0 = min(p[1] for p in box)
                    x1 = max(p[0] for p in box)
                    y1 = max(p[1] for p in box)
                    words.append(OCRWord(text=text, x0=x0, y0=y0, x1=x1, y1=y1))
            return words
        except Exception as e:
            logger.warning(f"PaddleOCR failed for {pdf_path} page {page_num}: {e}")
            return []


class EntityMatcher:
    def __init__(self, similarity_threshold: float = 0.75):
        self.sim_threshold = similarity_threshold

    def find_entity_spans(
        self,
        ocr_page: OCRPage,
        entity_name: str,
        values: List[str],
    ) -> List[EntitySpan]:
        if not values:
            return []

        spans: List[EntitySpan] = []
        ocr_text = [w.cleaned_text() for w in ocr_page.words]
        full_text = " ".join(ocr_text)

        for value in values:
            if not value:
                continue
            value_clean = self._clean_for_matching(value)

            # Strategy 1: Exact substring (case-insensitive)
            found = self._find_exact(full_text, ocr_text, entity_name, value_clean)
            if found:
                spans.extend(found)
                continue

            # Strategy 2: Substring of OCR text appears in value
            found = self._find_contained(full_text, ocr_text, entity_name, value_clean)
            if found:
                spans.extend(found)
                continue

            # Strategy 3: Fuzzy match (sliding window)
            found = self._find_fuzzy(ocr_text, entity_name, value_clean)
            if found:
                spans.extend(found)

        # Deduplicate overlapping spans (keep highest similarity)
        spans = self._deduplicate_spans(spans)
        return sorted(spans, key=lambda s: (-s.similarity, s.start_word_idx))

    def _clean_for_matching(self, text: str) -> str:
        t = text.upper()
        t = re.sub(r"[\.\,\;\:]+", " ", t)  # Remove punctuation
        t = re.sub(r"\s+", " ", t).strip()
        return t

    def _get_word_indices(self, full_text: str, ocr_text: List[str],
                           target: str) -> List[int]:
        # Rebuild word list for position mapping
        word_positions = []
        pos = 0
        for w in ocr_text:
            word_positions.append(pos)
            pos += len(w) + 1  # +1 for space

        target_lower = target.lower()
        full_lower = full_text.lower()
        start = 0
        indices = []
        while True:
            idx = full_lower.find(target_lower, start)
            if idx == -1:
                break
            # Convert character position to word index
            word_idx = self._char_pos_to_word(word_positions, idx)
            end_word_idx = self._char_pos_to_word(word_positions, idx + len(target))
            indices.append((word_idx, end_word_idx))
            start = idx + 1
        return indices

    def _char_pos_to_word(self, word_positions: List[int], char_pos: int) -> int:
        for i, wp in enumerate(word_positions):
            if wp > char_pos:
                return max(0, i - 1)
        return len(word_positions) - 1

    def _find_exact(self, full_text: str, ocr_text: List[str],
                    entity_name: str, value: str) -> List[EntitySpan]:
        # Strategy 1: exact substring match.
        spans = []
        indices = self._get_word_indices(full_text, ocr_text, value)
        for start, end in indices:
            matched = " ".join(ocr_text[start:end + 1])
            spans.append(EntitySpan(
                entity_name=entity_name,
                value=value,
                start_word_idx=start,
                end_word_idx=end,
                matched_text=matched,
                similarity=1.0,
            ))
        return spans

    def _find_contained(self, full_text: str, ocr_text: List[str],
                        entity_name: str, value: str) -> List[EntitySpan]:
        # Strategy 2: value contains long OCR words.
        spans = []
        value_lower = value.lower()
        words_in_value = value_lower.split()
        n = len(words_in_value)
        for i in range(len(ocr_text) - n + 1):
            window = " ".join(ocr_text[i:i + n]).lower()
            if window in value_lower:
                sim = SequenceMatcher(None, window, value_lower).ratio()
                if sim >= self.sim_threshold:
                    spans.append(EntitySpan(
                        entity_name=entity_name,
                        value=value,
                        start_word_idx=i,
                        end_word_idx=i + n - 1,
                        matched_text=" ".join(ocr_text[i:i + n]),
                        similarity=sim,
                    ))
        return spans

    def _find_fuzzy(self, ocr_text: List[str],
                    entity_name: str, value: str) -> List[EntitySpan]:
        # Strategy 3: sliding window fuzzy match.
        spans = []
        n_value_words = len(value.split())
        if n_value_words < 1:
            return spans

        best_sim = 0
        best_span: Optional[EntitySpan] = None

        for window_size in range(max(1, n_value_words - 2), n_value_words + 3):
            for i in range(len(ocr_text) - window_size + 1):
                window_text = " ".join(ocr_text[i:i + window_size])
                window_clean = self._clean_for_matching(window_text)
                sim = SequenceMatcher(None, window_clean, value).ratio()
                if sim > best_sim and sim >= self.sim_threshold:
                    best_sim = sim
                    best_span = EntitySpan(
                        entity_name=entity_name,
                        value=value,
                        start_word_idx=i,
                        end_word_idx=i + window_size - 1,
                        matched_text=window_text,
                        similarity=sim,
                    )
        if best_span:
            spans.append(best_span)
        return spans

    def _deduplicate_spans(self, spans: List[EntitySpan]) -> List[EntitySpan]:
        if not spans:
            return []
        spans = sorted(spans, key=lambda s: s.start_word_idx)
        result = []
        used = set()
        for span in spans:
            overlap = any(
                wi in used for wi in range(span.start_word_idx, span.end_word_idx + 1)
            )
            if not overlap:
                result.append(span)
                used.update(range(span.start_word_idx, span.end_word_idx + 1))
        return result
    

class LabeledDatasetBuilder:
    # Builds LayoutXLM training dataset from shipment folders.
    def __init__(
        self,
        similarity_threshold: float = 0.75,
        use_paddle_fallback: bool = True,
    ):
        self.reader = PDFOCRReader(use_paddle_fallback=use_paddle_fallback)
        self.matcher = EntityMatcher(similarity_threshold=similarity_threshold)

    def process_shipment(
        self,
        shipment_dir: Path,
    ) -> List[LabeledPage]:
        shipment_id = shipment_dir.name

        # Find Excel file
        excel_files = list(shipment_dir.glob("*.xlsx"))
        if not excel_files:
            logger.warning(f"No Excel in {shipment_id}")
            return []
        gt = GroundTruthReader(excel_files[0])
        all_entities = gt.get_all_entities()

        # Find PDF files
        pdf_types = {}
        for suffix, dtype in [("_CI.pdf", "CI"), ("_PL.pdf", "PL"), ("_BL.pdf", "BL")]:
            matches = list(shipment_dir.glob(f"*{suffix}"))
            if matches:
                pdf_types[dtype] = matches[0]

        labeled_pages: List[LabeledPage] = []

        for doc_type, pdf_path in pdf_types.items():
            try:
                ocr_pages = self.reader.read_pdf(pdf_path)
            except Exception as e:
                logger.warning(f"OCR failed for {pdf_path}: {e}")
                continue

            for ocr_page in ocr_pages:
                labeled = self._label_page(
                    shipment_id=shipment_id,
                    doc_type=doc_type,
                    ocr_page=ocr_page,
                    entities=all_entities,
                )
                if labeled:
                    labeled_pages.append(labeled)

        return labeled_pages

    def _label_page(
        self,
        shipment_id: str,
        doc_type: str,
        ocr_page: OCRPage,
        entities: Dict[str, List[str]],
    ) -> Optional[LabeledPage]:
        n_words = len(ocr_page.words)
        if n_words == 0:
            return None

        # Initialize all labels as O
        labels = [0] * n_words  # 0 = "O"
        all_spans: List[EntitySpan] = []

        # Find spans for each entity type
        for entity_name, values in entities.items():
            spans = self.matcher.find_entity_spans(ocr_page, entity_name, values)
            for span in spans:
                for wi in range(span.start_word_idx, span.end_word_idx + 1):
                    if wi < n_words:
                        if wi == span.start_word_idx:
                            labels[wi] = LABEL_TO_ID.get(f"B-{entity_name}", 0)
                        else:
                            labels[wi] = LABEL_TO_ID.get(f"I-{entity_name}", 0)
                all_spans.append(span)

        # Only return pages with at least one entity label
        if all(l == 0 for l in labels):
            return None

        return LabeledPage(
            shipment_id=shipment_id,
            doc_type=doc_type,
            page_num=ocr_page.page_num,
            width=ocr_page.width,
            height=ocr_page.height,
            words=ocr_page.words,
            labels=labels,
            spans=all_spans,
        )

    def to_layoutxlm_dict(self, page: LabeledPage) -> Dict[str, Any]:
        # Convert LabeledPage to LayoutXLM-compatible dict.
        return {
            "id": f"{page.shipment_id}_{page.doc_type}_p{page.page_num}",
            "text": [w.cleaned_text() for w in page.words],
            "bboxes": [w.bbox for w in page.words],
            "labels": page.labels,
            "width": page.width,
            "height": page.height,
            "doc_type": page.doc_type,
            "shipment_id": page.shipment_id,
        }


def summarize_dataset(pages: List[LabeledPage]) -> Dict[str, Any]:
    total_words = sum(len(p.labels) for p in pages)
    labeled_words = sum(1 for p in pages for l in p.labels if l != 0)
    total_spans = sum(len(p.spans) for p in pages)

    entity_counts: Dict[str, int] = defaultdict(int)
    for page in pages:
        for span in page.spans:
            entity_counts[span.entity_name] += 1

    return {
        "total_pages": len(pages),
        "total_words": total_words,
        "labeled_words": labeled_words,
        "label_coverage": labeled_words / max(total_words, 1),
        "total_spans": total_spans,
        "entity_counts": dict(sorted(entity_counts.items(), key=lambda x: -x[1])),
    }
